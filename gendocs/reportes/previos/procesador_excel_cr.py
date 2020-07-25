#!/usr/bin/env python3
# -*- coding: UTF-8 -*-

# Inmesoft v0.1
# Programó Daniel A. Esteban C. 3002927538

from openpyxl import load_workbook
import etiqueta_cr


class ProcesadorExcel():
    def __init__(self):
        filesheet = "libro_insumo_cr.xlsx"
        wb = load_workbook(filesheet)
        ws = wb.active
        datos = []
        fila = 2
        columna = 1
        while fila<=100:
            val = lambda c: ws.cell(row=fila, column=c).value
            ref = val(2)
            if ref == None:
                break
            tipo = val(3)
            cli = val(4)
            dest = val(5)
            plan = val(6)
            odp = val(7)
            dim = (val(13), val(14))
            hue = val(10)
            cal = val(12)
            gan = val(15)
            cant = val(16)
            
            estafila = [tipo, cli, dest, plan, odp, hue, cal, dim, gan, cant]
            datos.append(estafila)
            
            fila += 1

            self.datos = datos
            
    def __procesar_lista_etiquetas(self):
       
        lista_etiquetas = []
        for item in self.datos:
            [tipo, cli, dest, plan, odp, hue, cal, dim, gan, cant] = item
            # print("H:{} C:{} L:{} A:{}".format(hue, cal, dim[0], dim[1]))
            n = 1
            print("{} / {}".format(cli,plan))
            for n in range(cant):
                n += 1
                et =  [tipo, cli, dest, plan, odp, hue, cal, dim, gan, cant, n]
                print("H:{} C:{} L:{} A:{} N:{}/{}".format(hue, cal, dim[0], dim[1], n, cant))
                lista_etiquetas.append(et)
        return lista_etiquetas
                
    def generar_pdf(self):
        et = etiqueta_cr.PDFEtiqueta()
        lista_etiquetas = self.__procesar_lista_etiquetas()
        for item in lista_etiquetas:
            et.añadir_etiqueta(item)

        et.terminar()
if __name__ == "__main__":
    pe = ProcesadorExcel()
    pe.generar_pdf()