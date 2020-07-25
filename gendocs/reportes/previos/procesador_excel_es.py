#!/usr/bin/env python3
# -*- coding: UTF-8 -*-

# Inmesoft v0.1
# Programó Daniel A. Esteban C. 3002927538

from openpyxl import load_workbook
import etiqueta_es


class ProcesadorExcel():
    def __init__(self):

        filesheet = "libro_insumo_es.xlsx"
        wb = load_workbook(filesheet)
        ws = wb.active
        datos = []
        fila = 2
        columna = 1
        while fila<=300:
            val = lambda c: ws.cell(row=fila, column=c).value
            
            ref = val(2)
            if ref == None:
                break
            dim = (val(4)/100.0, val(3)/100.0)
            pls = (val(7), val(8), val(9), val(10))
            diam = (val(11), val(12))
            esp = (val(5), val(6))
            cant = val(18)
            peso = val(19)
            f = (3.141*7.86)/4000.0
            densidades = (round((float(diam[0])**2)*f,3), round((diam[1]**2)*f,3))
            # print(val(15), val(3), val(16), val(4))
            peso_malla = (val(15)*val(3)*densidades[0]/100.0 + val(16)*val(4)*densidades[1]/100.0)
            peso_tot = peso_malla*cant
            
            estafila = [ref, dim, diam, esp, pls, cant, peso_malla, peso_tot]
            datos.append(estafila)
            
            fila += 1
            
        # for d in datos:
            # print(d)
            self.datos = datos
            
    def __paquetes(self, cant, peso_und):
        peso_max = 1200
        cant_max = 40
        peso_tot = cant*peso_und
        l_paqs = []
        l_pesos = []
        exc_peso = (peso_tot > peso_max)
        exc_cant = (cant > cant_max)
        # print("Acomodar: {} mallas. {:.1f} kg/malla. {:.1f} kg en total.".format(cant, peso_und, peso_tot))
        if (not exc_peso) and (not exc_cant):
            lcant = cant
        elif exc_peso and exc_cant:
            mcant = int(peso_max/peso_und)
            if mcant > cant_max:
                lcant = cant_max
            else:
                lcant = mcant
        elif (not exc_peso) and exc_cant:
            # print("Cantidad")
            lcant = cant_max
        elif exc_peso and (not exc_cant):
            # print("Peso")
            lcant = int(peso_max/peso_und)
        coc, res = divmod(cant, lcant)
        hay_res = (not res == 0)
        ntp = coc + int(hay_res)
        # print("N. Paqs. T. {}".format(ntp))
        acum = 0
        for i in range(ntp-1):
            estepaq = cant//ntp
            l_paqs.append(estepaq)
            acum += estepaq
        ultimo = cant - acum
        l_paqs.append(ultimo)
        l_parejas = [(c, c*peso_und) for c in l_paqs]
        return l_parejas
        
    def __procesar_lista_etiquetas(self):
       
        lista_etiquetas = []
        for item in self.datos:
            [ref, dim, diam, esp, pls, cant_tot, peso_und, peso_tot] = item
            print("Ref: {} Cant: {} Peso: {:.1f}".format(ref, cant_tot, peso_tot))
            paqs = self.__paquetes(cant_tot, peso_und)
            n = 0
            nt = len(paqs)
            for cant_paq, peso_paq in paqs:
                n += 1
                et = [ref, dim, diam, esp, pls, cant_paq, peso_paq]
                print("    Paq {}/{}: Cant:{} Peso:{:.1f}".format(n, nt, cant_paq, peso_paq))
                lista_etiquetas.append(et)
        return lista_etiquetas
                
    def generar_pdf(self):
        encabezado = ("VERTICES URBANOS", "SENDEROS TRAPICHE")
        et = etiqueta_es.PDFEtiqueta(encabezado)
        lista_etiquetas = self.__procesar_lista_etiquetas()
        for item in lista_etiquetas:
            et.añadir_etiqueta(item)

        et.terminar()
if __name__ == "__main__":
    pe = ProcesadorExcel()
    pe.generar_pdf()